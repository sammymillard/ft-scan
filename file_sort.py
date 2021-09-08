# S.Millard and C.Christiansen - Aug 2021

# input excel files made for FT-scan
# move files listed in excel file in column D into one folder together

# standard library
import argparse
import os
from shutil import copy

# requires `pip install` in shell
from openpyxl import load_workbook

# define function
def main(excel_filename, current_dir, output_dir):

    # bring in excel sheet
    book = load_workbook(excel_filename)
    #print(book.sheetnames)
    sheet1 = book['Sheet1']

    # find correct column
    file_locations = sheet1['D']

    # loop through column
    count = 0

    ## get current file location
    for location in file_locations[1:]:
        if location.value is not None:
            count = count+1
            #print(location.value)

            current_location = (
                current_dir +
                location.value.encode("cp1252").decode("utf8")
            )
            #print(current_location)

            ## check location exists
            if os.path.isfile(current_location) is True:

                ## copy file to new directory
                copy(current_location, output_dir)
            else:
                print("Location of file not found for:")
                print(current_location)

            # try:
            #     copy(current_location, output_dir)
            # except FileNotFoundError:
            #     print("Location of file not found for:")
            #     print(current_location)
            #     print(current_location.__class__)

    # finished!
    print(count)
    print("Finished!")


# organisation for passing arguments
if __name__ == "__main__": # this is used when you call script from the shell
    parser = argparse.ArgumentParser(
        description=(
            "Move files from zotero folder structure to folders for "
            "FT-screening"
        )
    )
    parser.add_argument(
        'excel',
        metavar='Excel input',
        type=str,
        help='Excel with locations of separated files'
    )
    parser.add_argument(
        "-c",
        "--current_dir",
        help="Directory - where the files are now",
        #default=None
    )
    parser.add_argument(
        "-o",
        "--output_dir",
        help="Directory - where to save the files",
        default=None
    )
    args = parser.parse_args()

    # calling the function
    main(
        excel_filename=args.excel,
        current_dir=args.current_dir,
        output_dir=args.output_dir
    )


